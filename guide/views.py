from django.shortcuts import render, redirect
from django.http import HttpResponse

from .models import FinancialAccounting, Auditing, ManagementAccounting, CompanyLaw, ServiceMarketing, RetailMarketing
from .filters import FinancialAccountingFilter, AuditingFilter, ManagementAccountingFilter, CompanyLawFilter, ServiceMarketingFilter, RetailMarketingFilter
import openpyxl
# Create your views here.


def home(request):
    return render(request, 'guide/home.html')


def financial_accounting(request):
    queryset = FinancialAccounting.objects.all().order_by('s_no')

    my_filter = FinancialAccountingFilter(request.GET, queryset=queryset)
    queryset = my_filter.qs
    context = {
        'queryset': queryset,
        'my_filter': my_filter,
    }
    return render(request, 'guide/financial_accounting.html', context)


def financial_accounting_questions(request):
    work_book = openpyxl.load_workbook("FINANCIAL_MANAGEMENT.xlsx")

    for page in work_book.sheetnames:
        sheet = work_book[page]
        print(sheet.title)

        if sheet.title == 'Questions':

            # cell1 = sheet.cell(row=2, column=1)
            # result = cell1.value
            # s_no = result[0:1]
            # question = result[2:].lstrip()
            # print(s_no)
            # print(question)
            for row in range(2, sheet.max_row+1):
                if sheet.cell(row, column=1) != None:
                    try:
                        cell1 = sheet.cell(row, column=1)
                        result = cell1.value
                        if result[0:1].isdigit():
                            s_no = result[0:2]
                            s_no = int(s_no.replace('.', '').lstrip())
                            questions = result[4:].lstrip().partition('\n')
                            question = questions[0]
                            print(f'{s_no} - {question}')

                            FinancialAccounting.objects.create(
                                s_no=s_no, question=question)

                    except Exception:
                        pass

    return redirect('financial_accounting')


def financial_accounting_answers(request):
    work_book = openpyxl.load_workbook("FINANCIAL_MANAGEMENT.xlsx")

    for page in work_book.sheetnames:
        sheet = work_book[page]
        print(sheet.title)
        if sheet.title == 'Answers':
            for row in range(2, sheet.max_row+1):
                if sheet.cell(row, column=1) != None:
                    try:
                        cell1 = sheet.cell(row, column=1)
                        result = cell1.value
                        if result[0:1].isdigit():
                            s_no = result[0:2]
                            s_no = int(s_no.replace('.', '').lstrip())
                            answer = result[4:].lstrip()
                            print(f'{s_no} - {answer[0]}')

                            question = FinancialAccounting.objects.get(
                                s_no=s_no)
                            question.answer = answer
                            question.save()

                    except Exception:
                        pass

    return redirect('financial_accounting')


def auditing(request):
    queryset = Auditing.objects.all()

    my_filter = AuditingFilter(request.GET, queryset=queryset)
    queryset = my_filter.qs
    context = {
        'queryset': queryset,
        'my_filter': my_filter,
    }
    return render(request, 'guide/auditing.html', context)


def auditing_questions(request):
    work_book = openpyxl.load_workbook("Auditing.xlsx")

    for page in work_book.sheetnames:
        sheet = work_book[page]
        print(sheet.title)

        if sheet.title == 'Questions':

            for row in range(2, sheet.max_row+1):
                if sheet.cell(row, column=1) != None:
                    try:
                        cell1 = sheet.cell(row, column=1)
                        result = cell1.value
                        if result[0:1].isdigit():
                            s_no = result[0:2]
                            questions = result[4:].lstrip().partition('\n')
                            question = questions[0]
                            print(f'{s_no} - {question}')

                            Auditing.objects.create(
                                s_no=s_no, question=question)

                    except Exception:
                        pass

    return redirect('auditing')


def auditing_answers(request):
    work_book = openpyxl.load_workbook("Auditing.xlsx")

    for page in work_book.sheetnames:
        sheet = work_book[page]
        print(sheet.title)
        if sheet.title == 'Answers':
            for row in range(2, sheet.max_row+1):
                if sheet.cell(row, column=1) != None:
                    try:
                        cell1 = sheet.cell(row, column=1)
                        result = cell1.value
                        if result[0:1].isdigit():
                            s_no = result[0:2]
                            answer = result[4:].lstrip()
                            print(f'{s_no} - {answer[0]}')

                            question = Auditing.objects.get(
                                s_no=s_no)
                            question.answer = answer
                            question.save()

                    except Exception:
                        pass

    return redirect('auditing')


def management_accounting(request):
    queryset = ManagementAccounting.objects.all()

    my_filter = ManagementAccountingFilter(request.GET, queryset=queryset)
    queryset = my_filter.qs
    context = {
        'queryset': queryset,
        'my_filter': my_filter,
    }
    return render(request, 'guide/management_accounting.html', context)


def management_accounting_questions(request):
    work_book = openpyxl.load_workbook("Management_Accounting.xlsx")

    for page in work_book.sheetnames:
        sheet = work_book[page]
        print(sheet.title)

        if sheet.title == 'Questions':

            for row in range(2, sheet.max_row+1):
                if sheet.cell(row, column=1) != None:
                    try:
                        cell1 = sheet.cell(row, column=1)
                        result = cell1.value
                        if result[0:1].isdigit():
                            s_no = result[0:2]
                            questions = result[3:].lstrip().partition('\n')
                            question = questions[0]
                            print(f'{s_no} - {question}')

                            ManagementAccounting.objects.create(
                                s_no=s_no, question=question)

                    except Exception:
                        pass

    return redirect('management_accounting')


def management_accounting_answers(request):
    work_book = openpyxl.load_workbook("Management_Accounting.xlsx")

    for page in work_book.sheetnames:
        sheet = work_book[page]
        print(sheet.title)
        if sheet.title == 'Answers':
            for row in range(2, sheet.max_row+1):
                if sheet.cell(row, column=1) != None:
                    try:
                        cell1 = sheet.cell(row, column=1)
                        result = cell1.value
                        if result[0:1].isdigit():
                            s_no = result[0:2]
                            answer = result[4:].lstrip()
                            print(f'{s_no} - {answer[0]}')

                            question = ManagementAccounting.objects.get(
                                s_no=s_no)
                            question.answer = answer
                            question.save()

                    except Exception:
                        pass

    return redirect('management_accounting')


def company_law(request):
    queryset = CompanyLaw.objects.all()

    my_filter = CompanyLawFilter(request.GET, queryset=queryset)
    queryset = my_filter.qs
    context = {
        'queryset': queryset,
        'my_filter': my_filter,
    }
    return render(request, 'guide/company_law.html', context)


def company_law_questions(request):
    work_book = openpyxl.load_workbook("Company_Law.xlsx")

    for page in work_book.sheetnames:
        sheet = work_book[page]
        print(sheet.title)

        if sheet.title == 'Questions':

            for row in range(2, sheet.max_row+1):
                if sheet.cell(row, column=1) != None:
                    try:
                        cell1 = sheet.cell(row, column=1)
                        result = cell1.value
                        if result[0:1].isdigit():
                            s_no = result[0:2]
                            questions = result[3:].lstrip().partition('\n')
                            question = questions[0]
                            print(f'{s_no} - {question}')

                            CompanyLaw.objects.create(
                                s_no=s_no, question=question)

                    except Exception:
                        pass

    return redirect('company_law')


def company_law_answers(request):
    work_book = openpyxl.load_workbook("Company_Law.xlsx")

    for page in work_book.sheetnames:
        sheet = work_book[page]
        print(sheet.title)
        if sheet.title == 'Answers':
            for row in range(2, sheet.max_row+1):
                if sheet.cell(row, column=1) != None:
                    try:
                        cell1 = sheet.cell(row, column=1)
                        result = cell1.value
                        if result[0:1].isdigit():
                            s_no = result[0:2]
                            answer = result[4:].lstrip()
                            print(f'{s_no} - {answer[0]}')

                            question = CompanyLaw.objects.get(
                                s_no=s_no)
                            question.answer = answer
                            question.save()

                    except Exception:
                        pass

    return redirect('company_law')


def service_marketing(request):
    queryset = ServiceMarketing.objects.all()

    my_filter = ServiceMarketingFilter(request.GET, queryset=queryset)
    queryset = my_filter.qs
    context = {
        'queryset': queryset,
        'my_filter': my_filter,
    }
    return render(request, 'guide/service_marketing.html', context)


def service_marketing_questions(request):
    work_book = openpyxl.load_workbook("Service_Marketing.xlsx")

    for page in work_book.sheetnames:
        sheet = work_book[page]
        print(sheet.title)

        if sheet.title == 'Questions':

            for row in range(2, sheet.max_row+1):
                if sheet.cell(row, column=1) != None:
                    try:
                        cell1 = sheet.cell(row, column=1)
                        result = cell1.value
                        if result[0:1].isdigit():
                            s_no = result[0:2]
                            questions = result[3:].lstrip().partition('\n')
                            question = questions[0]
                            print(f'{s_no} - {question}')

                            ServiceMarketing.objects.create(
                                s_no=s_no, question=question)

                    except Exception:
                        pass

    return redirect('service_marketing')


def service_marketing_answers(request):
    work_book = openpyxl.load_workbook("Service_Marketing.xlsx")

    for page in work_book.sheetnames:
        sheet = work_book[page]
        print(sheet.title)
        if sheet.title == 'Answers':
            for row in range(2, sheet.max_row+1):
                if sheet.cell(row, column=1) != None:
                    try:
                        cell1 = sheet.cell(row, column=1)
                        result = cell1.value
                        if result[0:1].isdigit():
                            s_no = result[0:2]
                            answer = result[4:].lstrip()
                            print(f'{s_no} - {answer[0]}')

                            question = ServiceMarketing.objects.get(
                                s_no=s_no)
                            question.answer = answer
                            question.save()

                    except Exception:
                        pass

    return redirect('service_marketing')


def retail_marketing(request):
    queryset = RetailMarketing.objects.all()

    my_filter = RetailMarketingFilter(request.GET, queryset=queryset)
    queryset = my_filter.qs
    context = {
        'queryset': queryset,
        'my_filter': my_filter,
    }
    return render(request, 'guide/retail_marketing.html', context)


def retail_marketing_questions(request):
    work_book = openpyxl.load_workbook("Retail_Marketing.xlsx")

    for page in work_book.sheetnames:
        sheet = work_book[page]
        print(sheet.title)

        if sheet.title == 'Questions':

            for row in range(2, sheet.max_row+1):
                if sheet.cell(row, column=1) != None:
                    try:
                        cell1 = sheet.cell(row, column=1)
                        result = cell1.value
                        if result[0:1].isdigit():
                            s_no = result[0:2]
                            questions = result[3:].lstrip().partition('\n')
                            question = questions[0]
                            print(f'{s_no} - {question}')

                            RetailMarketing.objects.create(
                                s_no=s_no, question=question)

                    except Exception:
                        pass

    return redirect('retail_marketing')


def retail_marketing_answers(request):
    work_book = openpyxl.load_workbook("Retail_Marketing.xlsx")

    for page in work_book.sheetnames:
        sheet = work_book[page]
        print(sheet.title)
        if sheet.title == 'Answers':
            for row in range(2, sheet.max_row+1):
                if sheet.cell(row, column=1) != None:
                    try:
                        cell1 = sheet.cell(row, column=1)
                        result = cell1.value
                        if result[0:1].isdigit():
                            s_no = result[0:2]
                            answer = result[4:].lstrip()
                            print(f'{s_no} - {answer[0]}')

                            question = RetailMarketing.objects.get(
                                s_no=s_no)
                            question.answer = answer
                            question.save()

                    except Exception:
                        pass

    return redirect('retail_marketing')